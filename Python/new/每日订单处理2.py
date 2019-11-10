from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl,pprint
#wb = load_workbook("C:/Users/94960/Desktop/"+input("请输入文件名"))
#订单列表 (2).xlsx
wb = load_workbook("C:/Users/94960/Desktop/订单列表 (2).xlsx")
sheet=wb["列表1"]
courseDate = {}
#查询课程及其价格相关
for row in range(2,sheet.max_row+1):
    courseName=sheet["c"+str(row)].value #课程名
    coursePrice=sheet["E"+str(row)].value #课程价格
    courseday=sheet["M"+str(row)].value #课程时间
    #设置字典
    courseDate.setdefault(courseName,{
        "courseName": "xx",
        "buy_num":0,
        "coursePrice":0
    })
    #建立字典存储
    courseDate[courseName]["buy_num"] += 1
    courseDate[courseName]["coursePrice"] += float(coursePrice)
    courseDate[courseName]["courseName"] =courseName
    #course_number["courseName"]=courseName #建立课程名字典


"""length_jianzhidui=len(courseDate)
print("键值对数量%s个" % length_jianzhidui)
#输出数据
print("writing results...")
resultfile= open("每日处理后结果.py","w")
resultfile.write("alldata ="+pprint.pformat(courseDate))
resultfile.write("\n")
resultfile.write("alldata ="+pprint.pformat(course_number))"""


#建表  设置表头
ws1 = wb.create_sheet("results")
col=1
row=2
ws1.cell(1, 1,"课程名_群号")
ws1.cell(1, 2,"购买人数")
ws1.cell(1, 3,"价格总和")

def length_jianzhidui(args):
    pass
    buy_num1 = 1
    coursePrice1 = 1

#遍历字典，填数
buy_num1 = "1"
coursePrice1 = 1
for key in courseDate.keys():
    print(key)
    for key1 in courseDate[courseName].keys():
        if key1 == "courseName":
            courseName1 = courseName
            ws1.cell(column=col, row=row, value=key)
            col+=1

            buy_num1 = courseDate[key]["buy_num"]
            ws1.cell(column=col, row=row, value=buy_num1)
            col+=1

            coursePrice1 = courseDate[key]["coursePrice"]
            ws1.cell(column=col, row=row, value=coursePrice1)
            col=1
    row+=1

wb.save("C:/Users/94960/Desktop/订单列表 (2).xlsx")

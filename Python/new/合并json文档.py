#!/usr/bin/env python
# coding=utf-8
# json转换为excel
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl, pprint
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
import win32com.client as win32
import xlrd
import json
import os
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

cols = []

#filepath="C:\\Users\\94960\\Desktop\\python执行\\合并json\\小雅02 好友列表.json"
def json2excel(filepath,save_path,creat_sheet_choice,row_num_start):
    with open(filepath,'rb') as load_f:
        load_dict = json.load(load_f)
        #print(load_dict)
        #print("**********"*5)
        #print(type(load_dict[1]))
        #print(len(load_dict))
    len_load_dict=len(load_dict)

    #ws1 = wb.create_sheet("results")
    #sheet = wb.active

    if creat_sheet_choice == 1:
        ws1 = wb.create_sheet("results")
        sheet = wb.active
        col = 1
        row = 2
        ws1.cell(1, 1, "head_img")
        ws1.cell(1, 2, "nick_name")
        ws1.cell(1, 3, "remark_name")
        ws1.cell(1, 4, "user_name")
        ws1.cell(1, 5, "wxid")
    else :
        #ws1 = load_workbook(save_path)
        #print(type(ws1))
        #sheet = wb.active
        ws1=load_workbook(save_path)["results"]
        sheet=load_workbook(save_path)["results"]
        #print(sheet)
        row = row_num_start
        print("从%d行开始了"%row)
        col = 1
        ws1.cell(row, 1, "shuru")
    #操作字典
    for i in range(1,len_load_dict):
        dict_json=load_dict[i]
        for key_words in dict_json.keys():
            if key_words == "head_img":
                head_img = dict_json.get("head_img")
                ws1.cell(column=col, row=row, value=head_img)
                col += 1

                nick_name = dict_json.get("nick_name")
                ws1.cell(column=col, row=row, value=nick_name)
                col += 1

                remark_name = dict_json.get("remark_name")
                ws1.cell(column=col, row=row, value=remark_name)
                col += 1

                user_name = dict_json.get("user_name")
                ws1.cell(column=col, row=row, value=user_name)
                col += 1

                wxid = dict_json.get("wxid")
                ws1.cell(column=col, row=row, value=wxid)
            col = 1
        row += 1

    row_num=row-1
    print("result 表格一共有%d行" % row_num)

    # 判定result工作表的列宽，(实际需要加1奥),便于后面给表头上色加格式
    length_col_charm = 1
    for col_1 in range(1, ws1.max_column + 1):
        if sheet.cell(row=1, column=col_1).value is not None:  # 如果该单元格不为空
            length_col_charm = col_1
        """charm_title1=ws1.cell(row=1, column=col_1).value
        if charm_title1==" ":
            length_col_charm=col_1
            print("我是{}",col_1)
            break"""

    # 表格格式修改（行高、列宽、颜色、字号）
    ws1.column_dimensions['A'].width = 25.0  # 设置第一列宽
    ws1.row_dimensions[1].height = 16.0  # 设置第一行高
    # 设置格式
    boldBlackFont = Font(name='Times New Roman', bold=True, color=colors.WHITE)
    fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
    for col_num in range(1, length_col_charm + 1):
        ws1.cell(row=1, column=col_num).font = boldBlackFont
        ws1.cell(row=1, column=col_num).fill = fill
    # ws1['A1'] = 'Bold Red Times New Roman'
    row_num_start=row+2
    if creat_sheet_choice == 1:
        wb.save(save_path)
    else :
        load_workbook(save_path).save(save_path)
    return row_num_start
    #print(row_num_start)

"""
测试功能用
b="C:\\Users\\94960\\Desktop\\python执行\\合并json\\小雅02 好友列表.json"
a="C:/Users/94960/Desktop/数据输出.xlsx"
json2excel(b,a)
"""
output_path="C:/Users/94960/Desktop/数据输出.xlsx"
path = "C:\\Users\\94960\\Desktop\\python执行\\合并json1"
files= os.listdir(path)
creat_sheet_choice=1
row_num_start=1
allData = []
for file in files:
    if not os.path.isdir(file):
        print(file)
        input_path= path+"\\"+file
        print(input_path)
        b=json2excel(input_path,output_path,creat_sheet_choice,row_num_start)
        creat_sheet_choice = creat_sheet_choice+1
        print(creat_sheet_choice)
        row_num_start = b+1
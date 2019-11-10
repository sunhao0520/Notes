from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl,pprint
wb = load_workbook("C:/Users/94960/Desktop/python执行/订单列表 ("+input("请输入文件后辍数字")+").xlsx")
#订单列表 (2).xlsx
#wb = load_workbook("C:/Users/94960/Desktop/python执行/订单列表 (4).xlsx")
sheet=wb["列表1"]
courseDate = {}
#查询课程及其价格相关
for row in range(2,sheet.max_row+1):
    str_courseName=sheet["c"+str(row)].value #课程名
    courseName=str_courseName.split("\n")[0]
    #print(courseName)
    coursePrice=sheet["E"+str(row)].value #课程价格
    str_courseday=sheet["M"+str(row)].value #课程时间字符串
    course_day=str(str_courseday[0:10]) #时间取前10位
    #设置字典
    courseDate.setdefault(course_day,{})
    courseDate[course_day].setdefault(courseName,{
        "courseName": "xx",
        "buy_num":0,
        "coursePrice":0
    })
    #建立字典存储
    courseDate[course_day][courseName]["buy_num"] += 1
    courseDate[course_day][courseName]["coursePrice"] += float(coursePrice)
    courseDate[course_day][courseName]["courseName"] =courseName
    #course_number["courseName"]=courseName #建立课程名字典


length_jianzhidui=len(courseDate)
print("键值对数量%s个" % length_jianzhidui)
#输出数据
print("writing results...")
resultfile= open("每日处理后结果.py","w")
resultfile.write("alldata ="+pprint.pformat(courseDate))
resultfile.write("\n")

#建表  设置表头
ws1 = wb.create_sheet("results")
col=1
row=2
ws1.cell(1, 1,"课程名_群号")
ws1.cell(1, 2,"购买人数")
ws1.cell(1, 3,"价格总和")
ws1.cell(1, 4,"购买时间")


#遍历字典，填数
coursePrice1 = 1
for key_words in courseDate.keys(): #遍历日期
    print(key_words)
    for key in courseDate[key_words].keys(): #遍历课程
        print(key)
        for key1 in courseDate[key_words][key].keys(): #遍历数值
            if key1 == "courseName":
                courseName1 = courseName
                ws1.cell(column=col, row=row, value=key)
                col+=1

                #buy_num1 = courseDate[key_words][key]["buy_num"]
                buy_num1=courseDate[key_words][key].get("buy_num")
                ws1.cell(column=col, row=row, value=buy_num1)
                col+=1

                coursePrice1 = courseDate[key_words][key]["coursePrice"]
                ws1.cell(column=col, row=row, value=coursePrice1)
                col+=1

                course_day1=key_words
                ws1.cell(column=col, row=row, value=course_day1)
                col=1
        row+=1

wb.save("C:/Users/94960/Desktop/数据输出.xlsx")
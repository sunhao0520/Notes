from openpyxl import Workbook
from openpyxl import load_workbook

# 定位到工作簿
wb = load_workbook("C:/Users/94960/Desktop/test1.xlsx")
# 定位到表单
sheet = wb["lala"]

# 读取指定单元格
cell_values = sheet.cell(2, 3).value
print(cell_values)

# 获取最大的行数
row = sheet.max_row
# 获取最大列数
col = sheet.max_column
print(row, col)

# 数据计算
sheet.cell(5, 5).value = "=SUM(A5:D5)"
# 赋值给单元格
sheet.cell(6, 6).value = "%s,%s" % (row, col)
#修改后保存
wb.save("C:/Users/94960/Desktop/test1.xlsx")

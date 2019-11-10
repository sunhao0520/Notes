from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl,pprint
wb = load_workbook("C:/Users/94960/Desktop/"+input("请输入文件名"))
#订单列表 (2).xlsx
#wb = load_workbook("C:/Users/94960/Desktop/订单列表 (2).xlsx")
sheet=wb["列表1"]
courseDate = {}
course_number={}
#查询课程及其价格相关
for row in range(2,sheet.max_row+1):
    courseName=sheet["c"+str(row)].value #课程名
    coursePrice=sheet["E"+str(row)].value #课程价格
    #设置字典
    courseDate.setdefault(courseName,{
     #   "courseName1": "xx",
        "buy_num":0,
        "coursePrice":0
    })

    course_number.setdefault(courseName,0)
    #建立字典存储
    courseDate[courseName]["buy_num"] += 1
    courseDate[courseName]["coursePrice"] += float(coursePrice)
    #courseDate[courseName]["courseName1"] =courseName
    #course_number["courseName"]=courseName #建立课程名字典
    #课程数量

length_jianzhidui=len(courseDate)
print("键值对数量%s个" % length_jianzhidui)
#输出数据
print("writing results...")
resultfile= open("每日处理后结果.py","w")
resultfile.write("alldata ="+pprint.pformat(courseDate))

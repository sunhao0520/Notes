from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl,pprint
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
wb = load_workbook("C:/Users/94960/Desktop/python执行/订单列表 ("+input("请输入文件后辍数字")+").xlsx")
#订单列表 (2).xlsx
#wb = load_workbook("C:/Users/94960/Desktop/python执行/订单列表 (4).xlsx")
sheet=wb["列表1"]
courseDate = {}
i=0 #看计算次数的
title_state_id=9999 #给课程状态起始赋值

#查找需要的字段在表头的哪一列
for col in range(1,sheet.max_column+1):
    charm_title=sheet.cell(row=1, column=col).value
    if charm_title== "订单名称":
        title_name_id=get_column_letter(col)
        print("%s,%s"%(charm_title,title_name_id))
    elif charm_title== "实付金额":
        title_money_id=get_column_letter(col)
        print("%s,%s"%(charm_title,title_money_id))
    elif charm_title== "付款时间":
        title_time_id=get_column_letter(col)
        print("%s,%s"%(charm_title,title_time_id))
    elif charm_title== "状态":
        title_state_id=get_column_letter(col)
        print("%s,%s"%(charm_title,title_state_id))
    #print(charm_title)




#查询课程及其价格相关
for row in range(2,sheet.max_row+1):
    str_courseName=sheet[title_name_id+str(row)].value #课程名群号字符串
    courseName=str_courseName.split("\n")[0] #截取课程名
    #print(courseName)
    if title_state_id!=9999:
        buy_state = sheet[title_state_id + str(row)].value #判定课程状态一览是否存在，存在就输出，不存在就不输出
    else :
        buy_state ="absence"

    coursePrice=sheet[title_money_id+str(row)].value #课程价格（实付金额）

    str_courseday=str(sheet[title_time_id+str(row)].value) #课程时间字符串
    course_day=str(str_courseday[0:10]) #时间取前10位
    #course_day=sheet[title_time_id+str(row)].value
    #设置字典
    if buy_state=="已配送" or buy_state is None or buy_state=="absence": #判定订单配送状态(只执行已配送，
        # 缺失状态字段，或者状态值为空白的)
        courseDate.setdefault(course_day,{})
        courseDate[course_day].setdefault(courseName,{
            "courseName": "xx",
            "buy_num":0,
            "coursePrice":0
        })
        #建立字典存储
        courseDate[course_day][courseName]["buy_num"] += 1
        courseDate[course_day][courseName]["coursePrice"] += float(coursePrice)
        courseDate[course_day][courseName]["courseName"] =courseName
        #course_number["courseName"]=courseName #建立课程名字典
        i+=1 #计算一次，累加一次


length_jianzhidui=len(courseDate)
print("键值对数量%s个" % length_jianzhidui)
#输出数据到新的python文件里
print("writing results...")
resultfile= open("每日处理后结果.py","w")
resultfile.write("alldata ="+pprint.pformat(courseDate))
resultfile.write("\n")

#新建表单  设置表头
ws1 = wb.create_sheet("results")
col=1
row=2
ws1.cell(1, 1,"课程名称")
ws1.cell(1, 2,"购买人数")
ws1.cell(1, 3,"价格总和")
ws1.cell(1, 4,"购买时间")


#遍历字典，填数
coursePrice1 = 1
for key_words in courseDate.keys(): #遍历日期
    print(key_words)
    for key in courseDate[key_words].keys(): #遍历课程
        print(key)
        for key1 in courseDate[key_words][key].keys(): #遍历数值
            if key1 == "courseName":
                courseName1 = courseName
                ws1.cell(column=col, row=row, value=key)
                col+=1

                #buy_num1 = courseDate[key_words][key]["buy_num"]
                buy_num1=courseDate[key_words][key].get("buy_num")
                ws1.cell(column=col, row=row, value=buy_num1)
                col+=1

                coursePrice1 = courseDate[key_words][key]["coursePrice"]
                ws1.cell(column=col, row=row, value=coursePrice1)
                col+=1

                course_day1=key_words
                ws1.cell(column=col, row=row, value=course_day1)
                col=1
        row+=1

print(row)

#判定result工作表的列宽，(实际需要加1奥),便于后面给表头上色加格式
length_col_charm=1
for col_1 in range(1,ws1.max_column+1):
    if sheet.cell(row=1, column=col_1).value is not None: #如果该单元格不为空
        length_col_charm = col_1
    """charm_title1=ws1.cell(row=1, column=col_1).value
    if charm_title1==" ":
        length_col_charm=col_1
        print("我是{}",col_1)
        break"""

#表格格式修改（行高、列宽、颜色、字号）
ws1.column_dimensions['A'].width = 25.0 #设置第一列宽
ws1.row_dimensions[1].height = 16.0 #设置第一行高
#设置格式
boldBlackFont = Font(name='Times New Roman', bold=True, color= colors.WHITE)
fill = PatternFill(start_color ='800000', end_color = '800000', fill_type = 'solid')
for col_num in range(1,length_col_charm+1):
    ws1.cell(row=1,column=col_num).font = boldBlackFont
    ws1.cell(row=1, column=col_num).fill = fill
#ws1['A1'] = 'Bold Red Times New Roman'


print("*"*50)
print("一共%d个数据是有效订单，纳入了计算"%i)
wb.save("C:/Users/94960/Desktop/数据输出.xlsx")
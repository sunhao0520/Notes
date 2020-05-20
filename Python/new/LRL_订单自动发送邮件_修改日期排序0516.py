# coding=utf-8
import time, os, sched

##请在下面添加需要统计的课程编号号

xiugaicourse_input = ['C045', 'TP001', 'C079']
course_list = {
    "C045": "高阶写作",
    "C030": "万国西班牙语",
    "C079": "法语旁听训练营",
    "TP001": "观看之道"
}


#############

def Send_Dailymessage(keyIDs_input,course_list):
    # ************下面是邮件发送程序
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.image import MIMEImage
    from email.mime.application import MIMEApplication
    import datetime

    #############下面执行抓取数据，更新excel操作
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    import openpyxl, pprint
    from openpyxl.styles import Font
    from openpyxl.styles import colors
    from openpyxl.styles import PatternFill
    import win32com.client as win32
    import os
    import json
    import requests
    import datetime

    """name|require|means
    -|-|-
    productKeyId|0|产品编号，不传则为全部
    startDate|1|起始日期(包含该日期)，'2020-04-18'
    endDate|1|结束日期(包含该日期)，'2020-04-18'"""

    """num_startDate = input("请输入开始日期：") #这段是要用的
    num_endDate = input("请输入截止日期：")
    num_productKeyId=input("课程序号(不输入为总收入)：")"""

    """num_productKeyId="C0"+input("课程序号：")"""

    keyIDs = keyIDs_input

    # 遍历课程包
    wb = load_workbook("C:\\Users\\94960\\Desktop\\LRL数据汇总\\汇总.xlsx")
    sheet = wb["Money1"]
    sheet.delete_rows(1, sheet.max_row + 1)  # 清除原来的行数
    col = 1
    row = 2
    sheet.cell(1, 1, "日期")
    sheet.cell(1, 2, "课程编号")
    sheet.cell(1, 3, "支付宝金额")
    sheet.cell(1, 4, "微信支付金额")
    sheet.cell(1, 5, "CC订单有同步")
    sheet.cell(1, 6, "当日该课程收入")
    sheet.cell(1, 7, "备注：所有金额单位（元）")

    today = str(datetime.date.today())

    # 昨天的日期
    def getYesterday():
        today = datetime.date.today()
        oneday = datetime.timedelta(days=1)
        yesterday = today - oneday
        return yesterday

    yesterday = str(getYesterday())
    print(yesterday)
    print(type(yesterday))

    # 定义一个字典
    thatday = {}
    thatday_cunchu = []

    num_startDate = '2020-04-07'
    """num_startDate = "2020-"+input("请输入两位月份")+"-"+input("请输入两位日期")"""
    num_endDate = '2020-04-07'

    now = datetime.datetime.now()
    # 递减的时间
    delta = datetime.timedelta(days=-1)
    # 30天前的时间
    endnow = now + datetime.timedelta(days=-30)
    # 30天前的时间
    endnow = str(endnow.strftime('%Y-%m-%d'))

    offset = now

    # 当日期减为30天前的日期，循环结束
    while str(offset.strftime('%Y-%m-%d')) != endnow:
        offset += delta
        # print(str(offset.strftime('%Y-%m-%d')))

        num_startDate = str(offset.strftime('%Y-%m-%d'))
        num_endDate = str(offset.strftime('%Y-%m-%d'))

        IDzidian = {}
        for keyID in keyIDs:
            num_productKeyId = keyID

            web = "http://api.lrl.lonelyreader.com/tmp/countMoney?productKeyId=" + num_productKeyId + "&startDate=" + num_startDate + "&endDate=" + num_endDate
            # print(web)
            url = web
            r = requests.get(url)
            number = r.text

            zidian = json.loads(number)  # 字典
            # print(type(zidian))
            IDzidian.setdefault(keyID, {})
            IDzidian[keyID] = zidian
            # print(IDzidian)
            # print("*"*10)

            # 标记日期
        thatday_num = num_endDate
        # print(zidian[0]["totalFee"])
        thatday[thatday_num] = IDzidian

    shijianduan = num_startDate + "至" + num_endDate

    sheet.column_dimensions['A'].width = 15.0  # 设置第一列宽
    sheet.row_dimensions[1].height = 16.0  # 设置第一行高

    for day in thatday.keys():  # 遍历数值
        # print(day)

        # sheet.cell(column=1, row=row, value=day)
        # print(key)

        for course in thatday[day].keys():

            # print(course)
            # print(thatday[day][course])
            shuju = thatday[day][course]
            sheet.cell(column=2, row=row, value=course)
            sheet.cell(column=1, row=row, value=day)
            for item in shuju:  # 遍历数值
                #print(item)
                #print(day)
                if item is not None:
                    if item.get("type") == "ALI_PAY":
                        sheet.cell(column=3, row=row, value=("%.2f" % (item.get("totalFee") / 100)))

                    if item.get("type") == "WX_PAY":
                        sheet.cell(column=4, row=row, value=("%.2f" % (item.get("totalFee") / 100)))
                    if item.get("type") == "CC_SYNC_PAY":
                        sheet.cell(column=5, row=row, value="Yes")

                    total_money = "=C" + str(row) + "+" + "D" + str(row)
                    sheet.cell(column=6, row=row, value=total_money)
            row += 1
    pingtaishouru = "=SUM(F1" + ":" + "F" + str(row) + ")"
    sheet.cell(column=6, row=row + 4, value=pingtaishouru)
    sheet.cell(column=3, row=row + 4, value=("30天内课程总收入为"))

    # 为空值格子填入0
    i = 1
    j = 1
    for i in range(1, row):  # 对第1至row行单元格遍历（以下是对第一列的操作）
        for j in range(1, 7):
            if sheet.cell(row=i, column=j).value is None:  # 如果该单元格为空
                sheet.cell(i, j, 0)  # 那么填入n

    # 左对齐
    from openpyxl.styles import Alignment
    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for i in range(2, row):  # 对第1至row行单元格遍历（以下是对第一列的操作）
        for j in range(1, 7):
            # letter = get_column_letter(i + 1)
            sheet.cell(row=i, column=j).alignment = align
    # 判定result工作表的列数，(实际需要加1奥),便于后面给表头上色加格式
    length_col_charm = 1
    for col_1 in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col_1).value is not None:  # 如果该单元格不为空
            length_col_charm = col_1

    # 字体格式美化一下
    boldBlackFont = Font(bold=True, size=11)
    fill = PatternFill(start_color='FFD900', end_color='FFD900', fill_type='solid')
    for col_num in range(1, length_col_charm + 1):
        sheet.cell(row=1, column=col_num).font = boldBlackFont
        sheet.cell(row=1, column=col_num).fill = fill

    sheet.cell(row=1, column=7).font = Font(bold=True, size=9)
    sheet.column_dimensions['C'].width = 14.0  # 设置第3列宽
    sheet.column_dimensions['D'].width = 14.0  #
    sheet.column_dimensions['E'].width = 16.0  #
    sheet.column_dimensions['F'].width = 16.0  #
    sheet.column_dimensions['G'].width = 19.0  #
    #给各个课程加个名字
    for row in range(2, sheet.max_row + 1):
        str_courseID = sheet["B" + str(row)].value
        for key_course_id in course_list.keys():
            if str_courseID == key_course_id :
                sheet.cell(column=7, row=row, value=course_list[key_course_id])
                sheet.cell(column=7, row=row).font = Font(size=9)

    # 第二波计算开始
    ws1 = wb["Order2"]
    ws1.delete_rows(1, ws1.max_row + 1)  # 清除原来的行数

    col = 1
    row = 2
    ws1.cell(1, 1, "日期")
    ws1.cell(1, 2, "课程编号")
    ws1.cell(1, 3, "购买单价")
    ws1.cell(1, 4, "数量")
    ws1.cell(1, 5, "当日此单价下金额总和")
    ws1.cell(1, 6, "备注")

    now = datetime.datetime.now()
    # 递减的时间
    delta = datetime.timedelta(days=-1)
    # 30天前的时间
    endnow = now + datetime.timedelta(days=-30)
    # 30天前的时间
    endnow = str(endnow.strftime('%Y-%m-%d'))
    offset = now
    # 定义个新字典
    dailyorder = {}
    # 订单数量
    while str(offset.strftime('%Y-%m-%d')) != endnow:
        offset += delta
        # print(str(offset.strftime('%Y-%m-%d')))

        num_startDate = str(offset.strftime('%Y-%m-%d'))
        num_endDate = str(offset.strftime('%Y-%m-%d'))

        # 遍历课程包
        for keyID in keyIDs:
            num_productKeyId = keyID

            web1 = "http://api.lrl.lonelyreader.com/tmp/countOrder?productKeyId=" + num_productKeyId + "&startDate=" + num_startDate + "&endDate=" + num_endDate
            # print(web1)
            url1 = web1
            r1 = requests.get(url1)
            number_money = r1.text
            geshi = json.loads(number_money)
            #print (geshi)
            # print(type(geshi))
            # dailyorder[num_endDate] = geshi
            for item in geshi:
                if len(item) != 0:
                    ws1.cell(column=1, row=row, value=num_endDate)
                    ws1.cell(column=3, row=row, value=("%.2f" % (item.get("total_fee") / 100)))
                    ws1.cell(column=4, row=row, value=item.get("num"))
                    ws1.cell(column=2, row=row, value=num_productKeyId)
                    total_money = "=C" + str(row) + "*" + "D" + str(row)
                    ws1.cell(column=5, row=row, value=total_money)
                    row += 1

    order_money_sum = "=SUM(E1" + ":" + "E" + str(row) + ")"
    ws1.cell(column=5, row=row + 4, value=order_money_sum)
    ws1.cell(column=2, row=row + 4, value=("30天内的课程订单总金额为"))

    ws1.column_dimensions['A'].width = 15.0  # 设置第一列宽
    ws1.row_dimensions[1].height = 16.0  # 设置第一行高

    i = 1
    j = 1
    for i in range(1, row):  # 对第1至row行单元格遍历（以下是对第一列的操作）
        for j in range(1, 5):
            if ws1.cell(row=i, column=j).value is None:  # 如果该单元格为空
                ws1.cell(i, j, 0)  # 那么填入n
    from openpyxl.styles import Alignment
    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for i in range(2, row):  # 对第1至row行单元格遍历（以下是对第一列的操作）
        for j in range(1, 7):
            # letter = get_column_letter(i + 1)
            ws1.cell(row=i, column=j).alignment = align
    # 判定result工作表的列数，(实际需要加1奥),便于后面给表头上色加格式
    length_col_charm = 1
    for col_1 in range(1, ws1.max_column + 1):
        if ws1.cell(row=1, column=col_1).value is not None:  # 如果该单元格不为空
            length_col_charm = col_1

    # 字体格式美化一下
    boldBlackFont = Font(bold=True, size=10)
    fill = PatternFill(start_color='FFD900', end_color='FFD900', fill_type='solid')
    for col_num in range(1, length_col_charm + 1):
        ws1.cell(row=1, column=col_num).font = boldBlackFont
        ws1.cell(row=1, column=col_num).fill = fill

    row=1
    # 给各个课程加个名字
    for row in range(2, ws1.max_row + 1):
        str_courseID = ws1["B" + str(row)].value
        for key_course_id in course_list.keys():
            if str_courseID == key_course_id:
                ws1.cell(column=6, row=row, value=course_list[key_course_id])
                ws1.cell(column=6, row=row).font = Font(size=9)

    ws1.column_dimensions['E'].width = 19.0  #
    ws1.column_dimensions['F'].width = 16.0  #

    wb.save("C:\\Users\\94960\\Desktop\\LRL数据汇总\\汇总.xlsx")

    #############

    def getYesterday():
        today = datetime.date.today()
        oneday = datetime.timedelta(days=1)
        yesterday = today - oneday
        return yesterday

    yesterday = str(getYesterday())

    # 设置登录及服务器信息
    mail_host = 'smtp.126.com'
    mail_user = '18700593036'
    mail_pass = 'sunhao20080808'
    sender = '18700593036@126.com'
    receivers = ['sunhao@lonelyreader.com']

    # 设置eamil信息
    # 添加一个MIMEmultipart类，处理正文及附件
    message = MIMEMultipart()
    message['From'] = sender
    message['To'] = receivers[0]
    message['Subject'] = ("%s 订单汇总" % yesterday)

    part = MIMEText("你好，下面是LRL 高阶英文写作每日订单汇总")
    message.attach(part)

    # ---这是附件部分---
    # xlsx类型附件
    part = MIMEApplication(open('C:\\Users\\94960\\Desktop\\LRL数据汇总\\汇总.xlsx', 'rb').read())
    part.add_header('Content-Disposition', 'attachment', filename='LRL 近一月每日订单汇总.xlsx')
    message.attach(part)

    # 登录并发送
    try:
        smtpObj = smtplib.SMTP()
        smtpObj.connect(mail_host, 25)
        smtpObj.login(mail_user, mail_pass)
        smtpObj.sendmail(
            sender, receivers, message.as_string())
        print('message send success')
        smtpObj.quit()
    except smtplib.SMTPException as e:
        print('error', e)

    # ************上面是邮件发送程序


# 第一个参数确定任务的时间，返回从某个特定的时间到现在经历的秒数
# 第二个参数以某种人为的方式衡量时间
schedule = sched.scheduler(time.time, time.sleep)


def perform_command(cmd, inc):
    # 安排inc秒后再次运行自己，即周期运行
    schedule.enter(inc, 0, perform_command, (cmd, inc))
    os.system(cmd)


# 以需要的时间间隔执行某个命令

import time, os

# 以需要的时间间隔执行某个命令
now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))


def re_exe(cmd, inc=10):
    while True:
        os.system(cmd);
        Send_Dailymessage(xiugaicourse_input,course_list)
        print("%s运算了一次" % now)
        time.sleep(inc)


print("ready to excute scrapt...")
re_exe("echo hello world", 600)

# Send_Dailymessage(xiugaicourse_input)
